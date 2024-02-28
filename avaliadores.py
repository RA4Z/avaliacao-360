import openpyxl
import datetime

class Pendentes():
    def coletar_avaliacoes(username):
        currentDateTime = datetime.datetime.now()
        date = currentDateTime.date()
        year = date.strftime("%Y")

        # Abre o arquivo Excel
        workbook = openpyxl.load_workbook(f'Q:\GROUPS\BR_SC_JGS_WM_LOGISTICA\PCP\Robert\Vários\Avaliação\Avaliação {year}\Colaboradores.xlsm')

        # Seleciona a planilha desejada
        sheet = workbook['AvaliadorxAvaliado']

        avaliacoes_pendentes = []

        for row in sheet.iter_rows(min_row=1, max_col=5, max_row=sheet.max_row):
            valor_coluna_1 = row[0].value
            valor_coluna_2 = row[1].value
            status_atual = row[2].value
            cargo = row[4].value
            if valor_coluna_1 == username and (not status_atual):
                avaliacoes_pendentes.append({'colaborador':valor_coluna_2, 'cargo':cargo})

        # Fechar o arquivo após terminar de usar
        workbook.close()
        
        return avaliacoes_pendentes
    
    def concluir_avaliacao(avaliador, avaliado):
        currentDateTime = datetime.datetime.now()
        date = currentDateTime.date()
        year = date.strftime("%Y")

        # Abre o arquivo Excel
        workbook = openpyxl.load_workbook(f'Q:\GROUPS\BR_SC_JGS_WM_LOGISTICA\PCP\Robert\Vários\Avaliação\Avaliação {year}\Colaboradores.xlsm', keep_vba=True)

        # Seleciona a planilha desejada
        sheet = workbook['AvaliadorxAvaliado']

        for row in sheet.iter_rows(min_row=1, max_col=3, max_row=sheet.max_row):
            valor_coluna_1 = row[0].value
            valor_coluna_2 = row[1].value
            if valor_coluna_1 == avaliador and valor_coluna_2 == avaliado:
                row[2].value = "OK"

        workbook.save(f'Q:\GROUPS\BR_SC_JGS_WM_LOGISTICA\PCP\Robert\Vários\Avaliação\Avaliação {year}\Colaboradores.xlsm')
        workbook.close()
