import openpyxl
import datetime

class Questionario():
    def coletar_perguntas():
        currentDateTime = datetime.datetime.now()
        date = currentDateTime.date()
        year = date.strftime("%Y")

        # Abre o arquivo Excel
        workbook = openpyxl.load_workbook(f'Q:\GROUPS\BR_SC_JGS_WM_LOGISTICA\PCP\Robert\Vários\Avaliação\Avaliação {year}\Perguntas e Respostas.xlsx')

        # Seleciona a planilha desejada
        sheet = workbook['Perguntas']

        perguntas = []

        # Percorre as células da coluna 1 e coluna 2
        for row in sheet.iter_rows(min_row=2, max_col=6, max_row=sheet.max_row):
            numero = row[0].value
            pergunta = row[1].value
            pontos10 = row[2].value
            pontos8 = row[3].value
            pontos6 = row[4].value
            pontos4 = row[5].value
            perguntas.append({
                'numero': numero,
                'pergunta': pergunta,
                'pontos10': pontos10,
                'pontos8': pontos8,
                'pontos6': pontos6,
                'pontos4': pontos4,
                'score': 0
            })

        # Fechar o arquivo após terminar de usar
        workbook.close()
        
        return perguntas
    
    def coletar_cabecalhos():
        currentDateTime = datetime.datetime.now()
        date = currentDateTime.date()
        year = date.strftime("%Y")

        # Abre o arquivo Excel
        workbook = openpyxl.load_workbook(f'Q:\GROUPS\BR_SC_JGS_WM_LOGISTICA\PCP\Robert\Vários\Avaliação\Avaliação {year}\Perguntas e Respostas.xlsx')

        # Seleciona a planilha desejada
        sheet = workbook['Titulos']

        titulos = []

        # Percorre as células da coluna 1 e coluna 2
        for row in sheet.iter_rows(min_row=2, max_col=3, max_row=sheet.max_row):
            numero = row[0].value
            titulo = row[1].value
            topico = row[2].value
            titulos.append({
                'numero': numero,
                'titulo': titulo,
                'topico': topico
            })

        # Fechar o arquivo após terminar de usar
        workbook.close()
        
        return titulos

